#!/srv/python3/bin/python
"""
@Author ：kehongping
@Date   ：2024/4/30 10:31
@Desc   ：This is a tool for export mysql query result to excel file
"""
import argparse
import time
import sys
import logging

import xlwt
import openpyxl
import pymysql


def get_logger(name):
    logger = logging.getLogger(name)
    logger.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s %(name)s %(levelname)s: %(message)s')
    handler = logging.StreamHandler()
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    return logger


logger = get_logger('import_data')


def parse_options():
    parser = argparse.ArgumentParser(description='This is a tool for export mysql query result to excel file')
    parser.add_argument('-H', '--host', type=str, dest="host", required=False, default='127.0.0.1',
                        help="default mysql host: 127.0.0.1")
    parser.add_argument('-P', '--port', type=str, dest="port", required=False, default='3306',
                        help="default mysql port 3306")
    parser.add_argument('-u', '--user', type=str, dest="user", required=False, default='root',
                        help="default mysql user root")
    parser.add_argument('-p', '--password', type=str, dest="password", required=False, default='123456',
                        help="default mysql password 123456")
    parser.add_argument('-d', '--db', type=str, dest="db", required=True, default='', help="mysql db")
    parser.add_argument('-q', '--query', type=str, dest='query', required=False, help="mysql query")
    parser.add_argument('-o', '--output', type=str, dest='output', required=True, help="Output xls file name ")
    args = parser.parse_args()

    return args


def writeExcel(data, path):
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet

    for row_idx, row in enumerate(data):
        print(row)

    #     for col_idx, val in enumerate(row):
    #         outws.cell(row_idx +1, col_idx+1, val)
    #
    # outwb.save(path)  # 一定要记得保存


def get_msyql_query_result(host, port, user, password, db, query):
    """
    获取mysql查询结果
    :return:
    """
    try:
        conn = pymysql.connect(
            host=host,
            port=int(port),
            user=user,
            password=password,
            database=db,
            autocommit=True
        )
        logger.info("Successfully connected to MySQL database")

        with conn.cursor(pymysql.cursors.SSDictCursor) as cursor:
            # 执行 SQL 查询
            cursor.execute(query)

            result = cursor.fetchone()
            while result is not None:
                yield result
                result = cursor.fetchone()

    except Exception as e:
        logger.error(f"Error connecting to MySQL database: {e}")
    finally:
        cursor.close()
        conn.close()



if __name__ == "__main__":
    args = parse_options()
    start_time = time.time()

    if args.query:
        query = args.query
    else:
        query = sys.stdin.read().strip()
    writeExcel(get_msyql_query_result(args.host,args.port, args.user,args.password, args.db, query), args.output)
    end_time = time.time()
    logger.info(f"Export finish, cost time: {round(end_time - start_time, 2)}s")
